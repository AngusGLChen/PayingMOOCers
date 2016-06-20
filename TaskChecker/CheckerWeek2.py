'''
Created on Sep 29, 2015

@author: Angus
'''

import os
from openpyxl import load_workbook

def prepare_book(book):    
    rows = []    
    for sheet in book:
        for row in sheet.rows:
            rows.append(row)                        
    return rows

def ExerciseChecker(path):
    
    # Gather the solution
    solution_path = path + "Solutions/Week2.xlsx"
    solution = load_workbook(solution_path, read_only=True, use_iterators=False, 
                             keep_vba=False, guess_types=False, data_only=True)

    solution_rows = prepare_book(solution)
    num_solution_records = len(solution_rows)
    
    print "The number of solution records is: " + str(num_solution_records) + "\n"
    
    num_responses = 0
    num_correct_responses = 0
    
    # Output the checking result
    output_path = path + "Checking_results/week2_results"
    if os.path.isfile(output_path):
        os.remove(output_path)
    output = open(output_path, "w")    
    
    # Gather the response
    answer_path = path + "Response/Week_2/"
    files = os.listdir(answer_path)
    for file in files:
                
        file_type_array = file.split(".")
        file_type = file_type_array[len(file_type_array)-1]
        
        if file_type not in ["xlsx", "xlsm", "xltx", "xltm"]:
            continue
        
        #if file != "markus.xlsx":
        #    continue
        
        num_responses += 1
        
        check_result_array = []
        
        try:  
        
            answer = load_workbook(answer_path + file, read_only=True, use_iterators=False, 
                               keep_vba=False, guess_types=False, data_only=True)
        
            # Gather each sheet in the answer file
            for sheet in answer:
                                 
                answer_rows = {}
                num_answer_rows = 0
                       
                for row in sheet.rows:
                
                    num_answer_rows += 1
                    
                    person_name = str.lower(str.strip(str(row[0].value)))
                
                    remaining_elements = set()
                    for i in range(1, len(row)):
                        remaining_elements.add(str.lower(str.strip(str(row[i].value))))
                    answer_rows[person_name] = remaining_elements
            
                # Compare the answer and the solution
                check_result = True
                if num_answer_rows != 13738 and num_answer_rows != 13739:
                    check_result = False
                    check_result_array.append(check_result)
                    #print "False" + "\t" + "Number of rows\t" + str(num_answer_rows)
                else:
                    for row in solution_rows:
                        person_name = str.lower(str.strip(str(row[0].value)))
                        if not answer_rows.has_key(person_name):
                            check_result = False
                            #check_result_array.append(check_result)
                            #print "False" + "\t" + "Name\t" + person_name
                            break
                        else:
                            row_check_result = True
                        
                            for i in range(1, len(row)):
                                row_element = str.lower(str.strip(str(row[i].value)))
                                if row_element != "":
                                    if row_element not in answer_rows[person_name]:
                                        row_check_result = False
                                        check_result = False
                                        # check_result_array.append(check_result)
                                        # print "False" + "\t" + "Elements\t" + str(row_element)
                                        break
                            ''' 
                            solution_row_set =  set() 
                            for i in range(1, len(row)):
                                solution_row_set.add(str.lower(str(row[i].value)))
                            response_row_set = answer_rows[person_name]
                            for value in response_row_set:
                                if value != "" and value != "none" and value not in solution_row_set:
                                    # print person_name + "\t" + value
                                    print solution_row_set
                                    print response_row_set
                                    print 
                                    break
                            '''
                            
                                    
                                
                        
                            if not row_check_result:
                                break
                        
                    check_result_array.append(check_result)
                    
        except Exception as e:
            print "False\t" + str(e)
            check_result_array.append(False)
                        
        final_result = False
        for result in check_result_array:       
            if result:
                final_result = result
                num_correct_responses += 1
                break
        print file + "\t" + str(final_result)
        output.write(file + "\t" + str(final_result) + "\n")
        
    output.close()
    
    print "\n\n"
    print "The number of responses is: " + str(num_responses)
    print "The number of correct responses is: " + str(num_correct_responses)
    print "\n\n"
                        
                
            
            
     
         
                    
        
    
    
    
    




####################################################
path = "/Users/Angus/Data/EdX/EX101x/Bonus/Bonus Exercise/"
ExerciseChecker(path)
print "Finished."











