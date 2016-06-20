'''
Created on Sep 27, 2015

@author: Angus
'''

from collections import Counter
import os
from openpyxl import load_workbook
from pandas import DataFrame


def prepare_book(book):
    columns = []

    for sheet in book:
        for col in sheet.columns:
            cd = Counter([x.value for x in col])
            n_c = cd.most_common(1)[0]
            if n_c[1] > 100:
                del cd[n_c[0]]
            columns.append(cd)

    return columns


def column_similarity(a, b):
    try:
        diff = a-b
        la = len(a)
        lb = len(b)

        si = min(la, lb)/float(max(la, lb))
        dif = sum(diff.values())/float(max(sum(a.values()), sum(b.values())))

        return si*(1-dif)

    except:
        return 0


def estimate_quality(fi, r):
    ls = fi.split('.')

    if not ls[int(len(ls) - 1)] == 'xlsx':
        raise ValueError('Not an xlsx file')

    fn = os.path.join(r, fi)
    book = load_workbook(fn, read_only=True, use_iterators=False,
                        keep_vba=False, guess_types=False, data_only=True)
    prep = prepare_book(book)
    res = []

    for sc in columns:
        opt = 0
        for uc in prep:
            sim = column_similarity(sc, uc)
            if sim > opt:
                opt = sim
        res.append(opt)

    mr = {'submission': fi, 'quality': sum(res)/float(len(res))}

    return mr





num_sol = 200

res = {'submission': [], 'quality': [], 'week': []}

for rs,_,fs in [('/Users/Angus/Data/EdX/EX101x/Bonus/Bonus Exercise/Solutions/','',['Week_2.xlsx'])]: #os.walk('data//Solutions'):
        
    for fis in fs:
        week = fis.split('.')[0]
        
        fis = os.path.join(rs, fis)

        if fis.split('.')[1] != 'xlsx':
            continue

        count = 0
        solution = load_workbook(fis, read_only=True, use_iterators=False,
                                 keep_vba=False, guess_types=False, data_only=True)

        columns = prepare_book(solution)

        for r,_,f in os.walk('/Users/Angus/Data/EdX/EX101x/Bonus/Bonus Exercise/Response/Week_2/'):
            print r
            print f
            for fi in f:

                if count > num_sol:
                    break

                res['submission'].append(fi)
                res['week'].append(week.split('_')[1])

                try:
                    rm = estimate_quality(fi, r)
                    res['quality'].append(rm['quality'])
                    print(rm)
                except Exception as e:
                    res['quality'].append(-1)
                    print(e, 'fail')

                count += 1

df = DataFrame(res)
df.to_csv('results.csv')


