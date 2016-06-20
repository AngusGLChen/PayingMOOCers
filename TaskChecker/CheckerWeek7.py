'''
Created on Oct 1, 2015

@author: Angus
'''

import os
from openpyxl import load_workbook

def prepare_book(book):    
    rows = []    
    for sheet in book:
        for row in sheet.rows:
            rows.append(row)                        
    return rows

def ExerciseChecker(path):
    
    # Gather the solution
    solution_path = path + "Solutions/Week7.xlsx"
    solution = load_workbook(solution_path, read_only=True, use_iterators=False, 
                             keep_vba=False, guess_types=False, data_only=True)

    solution_rows = prepare_book(solution)
    num_solution_records = len(solution_rows)
    
    print "The number of solution records is: " + str(num_solution_records) + "\n"
    
    num_responses = 0
    num_correct_responses = 0
    
    # Output the checking result
    output_path = path + "Checking_results/week7_results"
    if os.path.isfile(output_path):
        os.remove(output_path)
    output = open(output_path, "w")    
    
    # Gather the response
    answer_path = path + "Response/Week_7/"
    files = os.listdir(answer_path)
    for file in files:
                        
        file_type_array = file.split(".")
        file_type = file_type_array[len(file_type_array)-1]
        
        if file_type not in ["xlsx", "xlsm", "xltx", "xltm"]:
            continue
        
        num_responses += 1
        check_result_array = []     
        
        try:
                        
            answer = load_workbook(answer_path + file, read_only=True, use_iterators=False, 
                               keep_vba=False, guess_types=False, data_only=True)
                
            # Gather each sheet in the answer file
            for sheet in answer:
                
                answer_rows = {}
                num_answer_rows = 0
                       
                for row in sheet.rows:
                    
                    id = row[0].value
                    num_answer_rows += 1
                
                    remaining_elements = set()
                    for i in range(1, len(row)):
                        remaining_elements.add(row[i].value)
                        
                    if id in answer_rows.keys():
                        answer_rows[id].append(remaining_elements)
                    else:
                        answer_rows[id] = [remaining_elements]
            
                # Compare the answer and the solution
                check_result = True                
                if num_answer_rows != 1289 and num_answer_rows != 1290:
                    check_result = False
                    check_result_array.append(check_result)                   
                else:                
                    for row in solution_rows:
                        id = row[0].value
                        if not answer_rows.has_key(id):
                            check_result = False
                            print "False\tID\t" + str(id)
                            break
                        else:
                            row_check_result = False
                        
                            from_value = row[1].value
                            to_value = row[2].value
                        
                            remaining_elements_array = answer_rows[id]
                            for remaining_elements in remaining_elements_array:
                                if from_value in remaining_elements and to_value in remaining_elements:
                                    row_check_result = True
                                    break
                                
                                    
                        
                            if not row_check_result:
                                # print "False\t" + str(id) + "\t" + str(from_value) + "\t" + str(to_value)
                                check_result = False
                                break                        
                        
                check_result_array.append(check_result)
                    
        except Exception as e:
            print "False\t" + str(e)
            check_result_array.append(False)            
                        
        final_result = False
        for result in check_result_array:       
            if result:
                final_result = result
                num_correct_responses += 1
                break
        print file + "\t" + str(final_result)
        output.write(file + "\t" + str(final_result) + "\n")
        
    output.close()
    
    print "\n\n"
    print "The number of responses is: " + str(num_responses)
    print "The number of correct responses is: " + str(num_correct_responses)
    print "\n\n"
                        
                
            
            
     
         
                    
        
    
    
    
    




####################################################
path = "/Users/Angus/Data/EdX/EX101x/Bonus/Bonus Exercise/"
ExerciseChecker(path)
print "Finished."