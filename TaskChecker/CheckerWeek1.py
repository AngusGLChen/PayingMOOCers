'''
Created on Sep 30, 2015

@author: Angus
'''

import os
from openpyxl import load_workbook

from pandas import DataFrame

def prepare_book(book):    
    rows = []    
    for sheet in book:
        for row in sheet.rows:
            rows.append(row)                        
    return rows

def ExerciseCheckerCorrect(path):
    
    # Gather the solution
    solution_path = path + "Solutions/Week1.xlsx"
    solution = load_workbook(solution_path, read_only=True, use_iterators=False, 
                             keep_vba=False, guess_types=False, data_only=True)

    solution_rows = prepare_book(solution)
    num_solution_records = len(solution_rows)
    
    print "The number of solution records is: " + str(num_solution_records) + "\n"
    
    num_responses = 0
    num_correct_responses = 0
    
    # Output the checking result
    output_path = path + "Checking_results/week1_results"
    if os.path.isfile(output_path):
        os.remove(output_path)
    output = open(output_path, "w")    
    
    # Gather the response
    answer_path = path + "Response/Week_1/"
    files = os.listdir(answer_path)
    for file in files:
                        
        file_type_array = file.split(".")
        file_type = file_type_array[len(file_type_array)-1]
        
        if file_type not in ["xlsx", "xlsm", "xltx", "xltm"]:
            continue
        
        num_responses += 1
        
        check_result_array = []   
        
        try:
                        
            answer = load_workbook(answer_path + file, read_only=True, use_iterators=False, 
                               keep_vba=False, guess_types=False, data_only=True)
        
            # Gather each sheet in the answer file
            for sheet in answer:
            
                answer_rows = {}
                num_answer_rows = 0
                       
                for row in sheet.rows:
                
                    num_answer_rows += 1
                    
                    full_address = str.lower(str.strip(str(row[0].value)))
                
                    remaining_elements = set()
                    for i in range(1, len(row)):
                        remaining_elements.add(str.lower(str.strip(str(row[i].value))))
                    answer_rows[full_address] = remaining_elements
            
                # Compare the answer and the solution
                check_result = True
                
                for row in solution_rows:
                    full_address = str.lower(str.strip(str(row[0].value)))
                    if not answer_rows.has_key(full_address):
                        check_result = False
                        break
                    else:
                        row_check_result = True
                        
                        for i in range(1, len(row)):
                            row_element = str.lower(str.strip(str(row[i].value)))
                            
                            if row_element not in answer_rows[full_address] and row_element[1:len(row_element)] not in answer_rows[full_address]:
                                                                       
                                row_check_result = False
                                check_result = False
                                #print "False" + "\t" + "Elements\t" + str(row_element) + "\t" + str(answer_rows[full_address])
                                break
                        
                        if not row_check_result:
                            break
                        
                check_result_array.append(check_result)
                    
        except Exception as e:
            print "False\t" + str(e)
            check_result_array.append(False)                   
                        
        final_result = False
        for result in check_result_array:       
            if result:
                final_result = result
                num_correct_responses += 1
                break
        #print file + "\t" + str(final_result)
        output.write(file + "\t" + str(final_result) + "\n")
        
    output.close()
    
    print "\n\n"
    print "The number of responses is: " + str(num_responses)
    print "The number of correct responses is: " + str(num_correct_responses)
    print "\n\n"


####################################################

def ExerciseCheckerAlmostCorrect(path):
    
    # Gather the solution
    solution_path = path + "Solutions/Week1.xlsx"
    solution = load_workbook(solution_path, read_only=True, use_iterators=False, 
                             keep_vba=False, guess_types=False, data_only=True)

    solution_rows = prepare_book(solution)
    num_solution_records = len(solution_rows)
    
    print "The number of solution records is: " + str(num_solution_records) + "\n"
    
    num_responses = 0
    
    all_accuracy_array = []
    almost_accuracy_array = []
    
    # Gather the response
    answer_path = path + "Response/Week_1/"
    files = os.listdir(answer_path)
    for file in files:
                        
        file_type_array = file.split(".")
        file_type = file_type_array[len(file_type_array)-1]
        
        if file_type not in ["xlsx", "xlsm", "xltx", "xltm"]:
            continue
        
        print file
        
        num_responses += 1
        
        num_check = 0
        num_contain = 0
         
        try:
                        
            answer = load_workbook(answer_path + file, read_only=True, use_iterators=False, 
                               keep_vba=False, guess_types=False, data_only=True)
        
            # Gather each sheet in the answer file
            for sheet in answer:
            
                answer_rows = {}
                       
                for row in sheet.rows:
                    full_address = str.lower(str.strip(str(row[0].value)))
                    remaining_elements = set()
                    for i in range(1, len(row)):
                        remaining_elements.add(str.lower(str.strip(str(row[i].value))))
                    answer_rows[full_address] = remaining_elements
            
                # Compare the answer and the solution
                for row in solution_rows:
                    full_address = str.lower(str.strip(str(row[0].value)))
                    if answer_rows.has_key(full_address):
                        for i in range(1, len(row)):
                            num_check += 1
                            row_element = str.lower(str.strip(str(row[i].value)))
                            
                            if row_element in answer_rows[full_address] or row_element[1:len(row_element)] in answer_rows[full_address]:
                                num_contain += 1
                    
        except Exception as e:
            print "False\t" + str(e)
                            
        if num_check > 0:            
            accuracy = float(num_contain) / num_check
            all_accuracy_array.append(accuracy)
            if accuracy < 1:
                almost_accuracy_array.append(accuracy)
        else:
            almost_accuracy_array.append(accuracy)
    
    print    
    df1 = DataFrame(all_accuracy_array)
    print "ALl: " + str(num_responses)
    print df1.mean()
    print df1.std()
    
    print 
    df2 = DataFrame(almost_accuracy_array)
    print "Almost: " + str(len(almost_accuracy_array))
    print df2.mean()
    print df2.std()

    
    




####################################################
path = "/Users/Angus/Data/EdX/EX101x/Bonus/Bonus Exercise/"
# ExerciseCheckerCorrect(path)
ExerciseCheckerAlmostCorrect(path)
print "Finished."












