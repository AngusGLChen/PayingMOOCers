'''
Created on Sep 30, 2015

@author: Angus
'''

import os
from openpyxl import load_workbook

def prepare_book(book):    
    columns = []    
    for sheet in book:
        for column in sheet.columns:
            columns.append(column)                        
    return columns

def ExerciseCheckerSorted(path):
    
    # Gather the solution
    solution_path = path + "Solutions/Week3.xlsx"
    solution = load_workbook(solution_path, read_only=True, use_iterators=False, 
                             keep_vba=False, guess_types=True, data_only=True)

    solution_columns = prepare_book(solution)    
            
    num_responses = 0
    num_correct_responses = 0
    
    # Output the checking result
    output_path = path + "Checking_results/week3_results"
    if os.path.isfile(output_path):
        os.remove(output_path)
    output = open(output_path, "w")    
    
    # Gather the response
    answer_path = path + "Response/Week_3/"
    files = os.listdir(answer_path)
    for file in files:
        
        file_array = file.split(".")
        
        file_id = file_array[0]             
        
        if file_id in ["0", "1"]:
            continue
        
        file_type = file_array[1]
        if file_type not in ["xlsx", "xlsm", "xltx", "xltm"]:
            continue
        
        num_responses += 1
        
        check_result_array = []  
        
        try:
            
            answer = load_workbook(answer_path + file, read_only=True, use_iterators=False, 
                               keep_vba=False, guess_types=False, data_only=True)
        
            # Gather each sheet in the answer file
            answer_columns = []  
            for sheet in answer:                       
                for answer_column in sheet.columns:                    
                    answer_column_array = []
                    for answer_element in answer_column:
                        if answer_element.value != None and answer_element.data_type == "n":
                            answer_column_array.append(round(answer_element.value, 4))       
                    answer_columns.append(answer_column_array)
        
            # Compare the answer and the solution            
            for column in solution_columns:
                
                check_result = True
                num_solution_element = len(column)
                
                solution_column = []
                for element in column:
                    solution_column.append(round(element.value, 4))                               
                
                # Comparison - Ascending order
                solution_column.sort(cmp=None, key=None, reverse=False)
                
                start_element_value = solution_column[0]
                for answer_column_array in answer_columns:                    
                    
                    if start_element_value in answer_column_array:                        
                        start_index = answer_column_array.index(start_element_value)                                     
                        
                        if start_index + num_solution_element - 1 <= len(answer_column_array):
                            for i in range(num_solution_element):
                                index = i + start_index                      
                            
                                if solution_column[i] != answer_column_array[index]:                                
                                    check_result = False
                                    break
                            
                                if i == num_solution_element - 1:
                                    check_result = True   
                                                         
                        if check_result:
                            break                    
                    else:
                        check_result = False
                        continue
                
                # Comparison - Descending order
                if not check_result:
                    solution_column.sort(cmp=None, key=None, reverse=True)
                                        
                    start_element_value = solution_column[0]
                    for answer_column_array in answer_columns:                    
                    
                        if start_element_value in answer_column_array:                  
                            start_index = answer_column_array.index(start_element_value)                        
                            
                            if start_index + num_solution_element - 1 <= len(answer_column_array):
                                for i in range(num_solution_element):
                                    index = i + start_index                                                
                            
                                    if solution_column[i] != answer_column_array[index]:                                
                                        check_result = False
                                        break
                            
                                    if i == num_solution_element - 1:
                                        check_result = True
                                                             
                            if check_result:
                                break                    
                        else:
                            check_result = False
                            continue
                        
                if check_result:    
                    check_result_array.append(check_result)
                    
        except Exception as e:
            print e
                     
        final_result = False
        if len(check_result_array) == len(solution_columns):
            final_result = True
            num_correct_responses += 1
        print file + "\t" + str(final_result)
        output.write(file + "\t" + str(final_result) + "\n")
        
    output.close()
    
    print "\n\n"
    print "The number of responses is: " + str(num_responses)
    print "The number of correct responses is: " + str(num_correct_responses)
    print "\n\n"
    
def ExerciseCheckerUnsorted(path):
    
    # Gather the solution
    solution_path = path + "Solutions/Week3.xlsx"
    solution = load_workbook(solution_path, read_only=True, use_iterators=False, 
                             keep_vba=False, guess_types=True, data_only=True)

    solution_columns = prepare_book(solution)    
            
    num_responses = 0
    num_correct_responses = 0
    
    # Output the checking result
    output_path = path + "Checking_results/week3_results_unsorted"
    if os.path.isfile(output_path):
        os.remove(output_path)
    output = open(output_path, "w")    
    
    # Gather the response
    answer_path = path + "Response/Week_3/"
    files = os.listdir(answer_path)
    for file in files:
        
        file_array = file.split(".")
        
        file_id = file_array[0]             
        
        if file_id in ["0", "1"]:
            continue
        
        file_type = file_array[1]
        if file_type not in ["xlsx", "xlsm", "xltx", "xltm"]:
            continue
        
        num_responses += 1
        
        check_result_array = []  
        
        try:
            
            answer = load_workbook(answer_path + file, read_only=True, use_iterators=False, 
                               keep_vba=False, guess_types=False, data_only=True)      
                                
            # Gather each sheet in the answer file
            answer_columns = []  
            for sheet in answer:                       
                for column in sheet.columns:
                    answer_columns.append(column)
        
            # Compare the answer and the solution
            for solution_column in solution_columns:
            
                for answer_column in answer_columns:
                    
                    num_correct_solution_element = 0                    
                    
                    answer_element_set = set()
                    for answer_element in answer_column:
                        if answer_element.value != None and answer_element.data_type == "n":                            
                            answer_element_set.add(round(answer_element.value, 4))                         
                    
                    solution_element_set = set()
                    for solution_element in solution_column:
                        solution_element_set.add(round(solution_element.value, 4))
                        
                    for solution_element in solution_element_set:
                        if solution_element in answer_element_set or str(solution_element) in answer_element_set:
                            num_correct_solution_element += 1
                        else:                            
                            break
                
                    if num_correct_solution_element == len(solution_element_set):
                        check_result_array.append(True)                        
                        break
                    
        except Exception as e:
            print e
                     
        final_result = False
        if len(check_result_array) == len(solution_columns):
            final_result = True
            num_correct_responses += 1
        print file + "\t" + str(final_result)
        output.write(file + "\t" + str(final_result) + "\n")
        
    output.close()
    
    print "\n\n"
    print "The number of responses is: " + str(num_responses)
    print "The number of correct responses is: " + str(num_correct_responses)
    print "\n\n"



####################################################
path = "/Users/Angus/Data/EdX/EX101x/Bonus/Bonus Exercise/"
#ExerciseCheckerSorted(path)
ExerciseCheckerUnsorted(path)
print "Finished."











